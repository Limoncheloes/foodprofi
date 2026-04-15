"""Tests for document generation service."""
import io
import pytest
from unittest.mock import MagicMock
import uuid
from datetime import datetime


def make_order(rest_name="Ресторан 1", user_name="Повар Иван"):
    order = MagicMock()
    order.id = uuid.uuid4()
    order.restaurant_name = rest_name
    order.user_name = user_name
    order.created_at = datetime(2026, 4, 14, 10, 30)
    return order


def make_item(name="Говядина", qty_ordered=10.0, qty_received=None,
              unit="кг", buyer_name="Айбек", cat_name="Мясо"):
    item = MagicMock()
    item.display_name = name
    item.quantity_ordered = qty_ordered
    item.quantity_received = qty_received
    item.unit = unit
    item.buyer_name = buyer_name
    item.category_name = cat_name
    item.raw_name = None
    item.is_catalog_item = True
    item.substitution_note = None
    return item


def test_generate_docx_returns_bytes():
    from app.services.documents import generate_docx
    order = make_order()
    items = [make_item(), make_item("Молоко", 5.0, unit="л", cat_name="Молочка")]
    result = generate_docx(order, items)
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_generate_docx_valid_word_file():
    """Output must be a valid .docx (ZIP) file."""
    import zipfile
    from app.services.documents import generate_docx
    order = make_order()
    items = [make_item()]
    result = generate_docx(order, items)
    buf = io.BytesIO(result)
    assert zipfile.is_zipfile(buf), "Output is not a valid ZIP/docx file"


def test_generate_xlsx_returns_bytes():
    from app.services.documents import generate_xlsx
    order = make_order()
    items = [make_item(), make_item("Картофель", 20.0, qty_received=18.5, unit="кг")]
    result = generate_xlsx(order, items)
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_generate_xlsx_valid_excel_file():
    """Output must be a valid .xlsx file readable by openpyxl."""
    import openpyxl
    from app.services.documents import generate_xlsx
    order = make_order()
    items = [make_item(qty_received=9.0)]
    result = generate_xlsx(order, items)
    wb = openpyxl.load_workbook(io.BytesIO(result))
    assert len(wb.sheetnames) >= 2  # Sheet1 = all items, Sheet2 = summary


def test_generate_xlsx_sheet1_has_correct_columns():
    import openpyxl
    from app.services.documents import generate_xlsx, XLSX_COLUMNS
    order = make_order()
    items = [make_item(qty_received=9.0)]
    result = generate_xlsx(order, items)
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    headers = [ws.cell(row=1, column=i).value for i in range(1, len(XLSX_COLUMNS) + 1)]
    assert headers == XLSX_COLUMNS


def test_generate_xlsx_data_row_correct():
    import openpyxl
    from app.services.documents import generate_xlsx
    order = make_order(rest_name="Ресторан Тест")
    item = make_item("Говядина", qty_ordered=10.0, qty_received=9.5, unit="кг")
    result = generate_xlsx(order, [item])
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    row = [ws.cell(row=2, column=i).value for i in range(1, 12)]
    assert "Ресторан Тест" in row
    assert "Говядина" in row
    assert 10.0 in row
    assert 9.5 in row
